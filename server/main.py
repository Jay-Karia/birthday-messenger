import os
import base64
import secrets
import asyncio
from datetime import datetime, date
from typing import Optional

from flask import Flask, request, jsonify
from flask_cors import CORS
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (
    Mail,
    Attachment,
    FileContent,
    FileName,
    FileType,
    Disposition,
)
from dotenv import load_dotenv
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook  # NEW

from components.whatsapp_msg import send_whatsapp
from components.text_ai import text_gen

# --- Load environment first ---
load_dotenv()

# --- App setup ---
app = Flask(__name__)
app.secret_key = os.getenv("APP_SECRET") or secrets.token_hex(16)

# CORS: electron renderer is file:// so we allow all origins (token auth still protects)
CORS(app, resources={r"/*": {"origins": "*"}})

# --- Auth / tokens ---
ACTIVE_TOKENS: set[str] = set()
AUTH_USER = os.getenv("AUTH_USER")
AUTH_PASS = os.getenv("AUTH_PASS")

# --- SendGrid / paths ---
SENDGRID_API_KEY = os.getenv("SENDGRID_API_KEY")
FROM_EMAIL = os.getenv("FROM_EMAIL", "devtest10292025@outlook.com")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CARD_TEMPLATE = os.path.join(BASE_DIR, "components", "card.png")
FONT_PATH = os.path.join(BASE_DIR, "components", "font-title.ttf")
EXCEL_PATH = os.path.normpath(os.path.join(BASE_DIR, "..", "data", "dummy.xlsx"))

# --- Utilities ---


def create_birthday_card(name: str, message: str) -> str:
    """Create a custom birthday card image with name & AI-generated message."""
    img = Image.open(CARD_TEMPLATE)
    draw = ImageDraw.Draw(img)

    name_font = ImageFont.truetype(FONT_PATH, 80)
    msg_font = ImageFont.truetype(FONT_PATH, 50)

    draw.text((200, 150), name, font=name_font, fill="blue")
    draw.text((200, 300), message, font=msg_font, fill="black")

    output_path = os.path.join(BASE_DIR, "birthday_card_custom.png")
    img.save(output_path)
    return output_path


def require_auth():
    """Check Bearer token, return None if ok else (response, code)."""
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header[7:].strip()
        if token in ACTIVE_TOKENS:
            return None
    return jsonify({"error": "Unauthorized"}), 401


def parse_month_day(date_str: str) -> Optional[str]:
    """
    Accept MM-DD or YYYY-MM-DD and return MM-DD or None if invalid.
    """
    if not date_str:
        return None
    try:
        if len(date_str) == 5 and date_str[2] == "-":
            datetime.strptime(f"2000-{date_str}", "%Y-%m-%d")
            return date_str
        if len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            return dt.strftime("%m-%d")
    except ValueError:
        return None
    return None


def _coerce_birthday(cell_val) -> Optional[datetime]:
    if isinstance(cell_val, datetime):
        return cell_val
    if isinstance(cell_val, date):
        return datetime.combine(cell_val, datetime.min.time())
    if isinstance(cell_val, str):
        s = cell_val.strip()
        # Try supported formats
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def read_excel_matches(month_day: str) -> list[dict]:
    """
    Read the Excel workbook (headerless or with a header row) and build rows
    that match month-day of birthday.
    """
    matches: list[dict] = []
    if not os.path.exists(EXCEL_PATH):
        return matches

    try:
        wb = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    except Exception as e:
        print(f"[read_excel_matches] Failed to open workbook: {e}")
        return matches

    # Use first sheet
    try:
        sheet = wb[wb.sheetnames[0]]
    except Exception:
        return matches

    for row in sheet.iter_rows(values_only=True):
        # Expect at least 4 columns to attempt parse
        if not row or len(row) < 4:
            continue

        birthday_raw = row[3]
        bday = _coerce_birthday(birthday_raw)
        if not bday:
            # Likely a header row or invalid entry
            continue

        if bday.strftime("%m-%d") == month_day:
            def _safe(idx):
                return row[idx] if len(row) > idx and row[idx] is not None else ""

            matches.append({
                "id": _safe(0),
                "name": _safe(1),
                "app_id": _safe(2),
                "birthday": bday.strftime("%Y-%m-%d"),
                "email": _safe(4),
                "phone": _safe(5),
                "father_email": _safe(6),
                "father_phone": _safe(7),
                "mother_email": _safe(8),
                "mother_phone": _safe(9),
            })

    return matches


# --- Routes ---


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"}), 200


@app.route("/login", methods=["POST"])
def login():
    data = request.json or {}
    user = data.get("user")
    password = data.get("password")

    if user == AUTH_USER and password == AUTH_PASS:
        token = secrets.token_urlsafe(32)
        ACTIVE_TOKENS.add(token)
        return jsonify({"message": "Login successful", "token": token}), 200

    return jsonify({"error": "Invalid credentials"}), 401


@app.route("/logout", methods=["POST"])
def logout():
    token = request.headers.get("Authorization", "").removeprefix("Bearer ").strip()
    if token in ACTIVE_TOKENS:
        ACTIVE_TOKENS.remove(token)
    return jsonify({"message": "Logged out"}), 200


@app.route("/filter", methods=["GET"])
def filter_birthdays():
    auth_error = require_auth()
    if auth_error:
        return auth_error

    date_str = request.args.get("date")
    month_day = parse_month_day(date_str)
    if not month_day:
        return jsonify({"error": "Invalid date format. Use MM-DD or YYYY-MM-DD"}), 400

    matches = read_excel_matches(month_day)
    formatted_date = datetime.strptime(f"2000-{month_day}", "%Y-%m-%d").strftime("%B %d")

    return jsonify({
        "date": formatted_date,
        "month_day": month_day,
        "count": len(matches),
        "people": matches
    }), 200


@app.route("/send_card", methods=["POST"])
def send_email():
    auth_error = require_auth()
    if auth_error:
        return auth_error

    data = request.json or {}
    recipient = data.get("recipient")
    if not recipient:
        return jsonify({"error": "Recipient email required"}), 400

    name = data.get("name", "Friend")
    subject = data.get("subject", f"Happy Birthday, {name} 🎂")
    recipient_phone = data.get("recipient_phone", "919486870915")

    father_email = data.get("father_email")
    father_phone = data.get("father_phone")
    mother_email = data.get("mother_email")
    mother_phone = data.get("mother_phone")

    extra_recipients = []
    if father_email:
        extra_recipients.append(father_email)
    if mother_email:
        extra_recipients.append(mother_email)
    extra_phones = []
    if father_phone:
        extra_phones.append(str(father_phone))
    if mother_phone:
        extra_phones.append(str(mother_phone))

    # Generate AI message
    try:
        message = asyncio.run(text_gen(name))
    except Exception as e:
        return jsonify({"error": f"AI text generation failed: {e}"}), 500

    # Build card
    try:
        card_path = create_birthday_card(name, message)
    except Exception as e:
        return jsonify({"error": f"Card generation failed: {e}"}), 500

    # Encode image
    try:
        with open(card_path, "rb") as f:
            encoded_file = base64.b64encode(f.read()).decode()
    except Exception as e:
        return jsonify({"error": f"Attachment encoding failed: {e}"}), 500

    attachment = Attachment(
        FileContent(encoded_file),
        FileName(os.path.basename(card_path)),
        FileType("image/png"),
        Disposition("attachment"),
    )

    all_recipients = [recipient] + extra_recipients

    # Send WhatsApp (best-effort)
    all_phones = [str(recipient_phone)] + extra_phones
    for phone in all_phones:
        try:
            send_whatsapp(message, phone, card_path)
        except Exception:
            pass

    try:
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        for email_addr in all_recipients:
            email = Mail(
                from_email=FROM_EMAIL,
                to_emails=email_addr,
                subject=subject,
                html_content=f"<p>{message}</p>",
            )
            email.attachment = attachment
            sg.send(email)
        return jsonify({
            "status": 200,
            "message": "Email sent successfully to all recipients!",
        }), 200
    except Exception as e:
        return jsonify({"error": f"Email send failed: {e}"}), 500


# --- Error Handlers ---


@app.errorhandler(404)
def not_found(_e):
    return jsonify({"error": "Not found"}), 404


@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": f"Server error: {e}"}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=True)